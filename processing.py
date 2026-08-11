from io import BytesIO

import openpyxl
import pandas as pd

COL = {
    "Year": 7,
    "Category": 10,
    "Bag No": 11,
    "Style No": 12,
    "Department": 14,
    "Rec Dept": 22,
    "Bag Type": 21,
    "NetWt (B)": 18,
    "Loss Wt": 19,
    "In Wt Gm": 29,
    "Out Wt Gm": 32,
    "Base Metal": 52,
    "Raw Name": 53,
}
IDX = {name: col - 1 for name, col in COL.items()}


def _cell(row_values, col_name, n):
    idx = IDX[col_name]
    return row_values[idx] if idx < n else None

REQUIRED_HEADERS = {"Bag No": 11, "Raw Name": 53, "Category": 10}
MAX_HEADER_SCAN_ROWS = 5


class InvalidWorkbookError(Exception):
    pass


def _row_matches_headers(row_values):
    matches = 0
    for name, idx in REQUIRED_HEADERS.items():
        header_value = row_values[idx - 1] if idx - 1 < len(row_values) else None
        if header_value is not None and name.split(" ")[0].lower() in str(header_value).lower():
            matches += 1
    return matches == len(REQUIRED_HEADERS)


def _find_data_sheet(wb):
    scanned = {}
    for ws in wb.worksheets:
        preview_row = ()
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN_ROWS, values_only=True), start=1):
            if _row_matches_headers(row):
                return ws, row_idx
            if not preview_row and any(v is not None for v in row):
                preview_row = row
        scanned[ws.title] = preview_row
    return None, scanned


def _load_workbook_state(uploaded_file, metal_raw_name):
    try:
        wb = openpyxl.load_workbook(BytesIO(uploaded_file.getvalue()), data_only=True, read_only=True)
    except Exception as exc:
        raise InvalidWorkbookError(
            f"Could not open '{uploaded_file.name}' as an Excel file. It may be corrupted, "
            f"password-protected, or not a real .xlsx file. ({exc})"
        ) from exc

    ws, header_row = _find_data_sheet(wb)
    if ws is None:
        sheet_summaries = "; ".join(f"'{name}' starts with {preview[:6]}" for name, preview in header_row.items())
        raise InvalidWorkbookError(
            f"None of the sheets in '{uploaded_file.name}' look like a Bag Transaction export "
            f"(expected columns like 'Bag No', 'Category', 'Raw Name' within the first "
            f"{MAX_HEADER_SCAN_ROWS} rows). Found sheets: {sheet_summaries}. This may be a different "
            "report (e.g. a bag/stock summary) rather than the transaction-detail export this app needs."
        )

    bag_data = {}
    order = []
    bag_running_balance = {}
    bag_last_wt = {}
    bag_label = {}
    stage_rows = []

    for row_values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        n = len(row_values)
        marker = row_values[0] if n > 0 else None
        if marker is not None and str(marker).startswith("Bag:"):
            continue

        bag_no = _cell(row_values, "Bag No", n)
        if bag_no is None:
            continue

        if bag_label.get(bag_no) is None or bag_label[bag_no] == str(bag_no):
            year = _cell(row_values, "Year", n)
            bag_type = _cell(row_values, "Bag Type", n)
            bag_label[bag_no] = f"{year}/{bag_type}/{bag_no}" if year and bag_type else str(bag_no)

        net_wt_any = _cell(row_values, "NetWt (B)", n)
        if net_wt_any not in (None, 0):
            bag_last_wt[bag_no] = net_wt_any

        raw_name = _cell(row_values, "Raw Name", n)
        if raw_name != metal_raw_name:
            continue

        category = _cell(row_values, "Category", n)
        department = _cell(row_values, "Department", n)
        net_wt = net_wt_any or 0
        loss_wt = _cell(row_values, "Loss Wt", n) or 0
        in_wt = _cell(row_values, "In Wt Gm", n) or 0
        out_wt = _cell(row_values, "Out Wt Gm", n) or 0

        if bag_no not in bag_data:
            opening = net_wt if net_wt else in_wt
            add_metal = 0.0
            return_metal = round(out_wt, 3)
            loss_metal = round(loss_wt, 3)
            balance = round(opening + add_metal - return_metal - loss_metal, 3)
            bag_data[bag_no] = {
                "Category": category,
                "Opening Wt": round(opening, 3),
                "GRN Net Wt + Loss Wt": None,
                "Add Metal": add_metal,
                "Return Metal": return_metal,
                "Loss Metal": loss_metal,
            }
            order.append(bag_no)
        else:
            prev_balance = bag_running_balance[bag_no]
            opening = prev_balance
            add_metal = round(in_wt, 3)
            return_metal = round(out_wt, 3)
            loss_metal = round(loss_wt, 3)
            balance = round(prev_balance + add_metal - return_metal - loss_metal, 3)
            d = bag_data[bag_no]
            d["Add Metal"] = round(d["Add Metal"] + add_metal, 3)
            d["Return Metal"] = round(d["Return Metal"] + return_metal, 3)
            d["Loss Metal"] = round(d["Loss Metal"] + loss_metal, 3)
            if category and not d["Category"]:
                d["Category"] = category

        if department == "GRN":
            bag_data[bag_no]["GRN Net Wt + Loss Wt"] = round(net_wt + loss_wt, 3)

        bag_running_balance[bag_no] = balance

        stage_rows.append(
            {
                "Bag No": bag_no,
                "Category": category,
                "Department": department,
                "Opening Wt": round(opening, 3),
                "Add Metal": add_metal,
                "Return Metal": return_metal,
                "Loss Metal": loss_metal,
                "Balance": balance,
            }
        )

    return {
        "bag_data": bag_data,
        "order": order,
        "bag_running_balance": bag_running_balance,
        "bag_last_wt": bag_last_wt,
        "bag_label": bag_label,
        "stage_rows": stage_rows,
    }


def _as_file_list(uploaded_files):
    if isinstance(uploaded_files, (list, tuple)):
        return list(uploaded_files)
    return [uploaded_files]


def _build_gross_loss_report_single(uploaded_file, metal_raw_name, mismatch_tolerance, exclude_recast):
    state = _load_workbook_state(uploaded_file, metal_raw_name)
    bag_data = state["bag_data"]
    bag_running_balance = state["bag_running_balance"]
    bag_last_wt = state["bag_last_wt"]
    bag_label = state["bag_label"]
    recast_bags = _find_returned_recast_bags(state) if exclude_recast else {}

    rows = []
    for bag_no in state["order"]:
        if bag_no in recast_bags:
            continue
        d = bag_data[bag_no]
        balance = bag_running_balance[bag_no]
        last_wt = bag_last_wt.get(bag_no)
        gross_loss_pct = round((d["Loss Metal"] / balance) * 100, 3) if balance else None

        if last_wt is None:
            status = "No Data"
        elif abs(balance - last_wt) <= mismatch_tolerance:
            status = "OK"
        else:
            status = "Mismatch"

        rows.append(
            {
                "Source File": uploaded_file.name,
                "Bag No": bag_label[bag_no],
                "Category": d["Category"],
                "Opening Wt": d["Opening Wt"],
                "GRN Net Wt + Loss Wt": d["GRN Net Wt + Loss Wt"],
                "Add Metal": d["Add Metal"],
                "Return Metal": d["Return Metal"],
                "Loss Metal": d["Loss Metal"],
                "Balance": balance,
                "Last Weight": last_wt,
                "Gross Loss %": gross_loss_pct,
                "Status": status,
            }
        )

    return rows


def build_gross_loss_report(uploaded_files, metal_raw_name="Gold", mismatch_tolerance=0.001, exclude_recast=False):
    all_rows = []
    for uploaded_file in _as_file_list(uploaded_files):
        all_rows.extend(
            _build_gross_loss_report_single(uploaded_file, metal_raw_name, mismatch_tolerance, exclude_recast)
        )
    return pd.DataFrame(all_rows)


def build_department_summary(uploaded_files, metal_raw_name="Gold", exclude_recast=False):
    all_stage_rows = []
    for uploaded_file in _as_file_list(uploaded_files):
        state = _load_workbook_state(uploaded_file, metal_raw_name)
        recast_bags = _find_returned_recast_bags(state) if exclude_recast else {}
        all_stage_rows.extend(row for row in state["stage_rows"] if row["Bag No"] not in recast_bags)

    stage_df = pd.DataFrame(all_stage_rows)
    if stage_df.empty:
        return stage_df

    summary = (
        stage_df.groupby("Department", dropna=False)
        .agg(
            Bags=("Bag No", "nunique"),
            Entries=("Bag No", "count"),
            **{
                "Total Opening Wt": ("Opening Wt", "sum"),
                "Total Add Metal": ("Add Metal", "sum"),
                "Total Return Metal": ("Return Metal", "sum"),
                "Total Loss Metal": ("Loss Metal", "sum"),
                "Total Balance": ("Balance", "sum"),
            },
        )
        .reset_index()
    )
    numeric_cols = summary.columns.drop(["Department", "Bags", "Entries"])
    summary[numeric_cols] = summary[numeric_cols].round(3)
    summary["Gross Loss %"] = (
        summary["Total Loss Metal"] / summary["Total Balance"].replace(0, pd.NA) * 100
    ).round(3)
    summary = summary.sort_values("Total Loss Metal", ascending=False)

    total_row = {col: None for col in summary.columns}
    total_row["Department"] = "TOTAL (sum of %)"
    total_row["Gross Loss %"] = round(summary["Gross Loss %"].sum(), 3)
    summary = pd.concat([summary, pd.DataFrame([total_row])], ignore_index=True)
    return summary


def build_category_weight_summary(df: pd.DataFrame) -> pd.DataFrame:
    summary = (
        df.groupby("Category", dropna=False)
        .agg(
            Bags=("Bag No", "count"),
            **{
                "Total Opening Wt": ("Opening Wt", "sum"),
                "Total Add Metal": ("Add Metal", "sum"),
                "Total Return Metal": ("Return Metal", "sum"),
                "Total Loss Metal": ("Loss Metal", "sum"),
                "Total Balance": ("Balance", "sum"),
                "Avg Gross Loss %": ("Gross Loss %", "mean"),
            },
        )
        .reset_index()
        .sort_values("Total Loss Metal", ascending=False)
    )
    numeric_cols = summary.columns.drop(["Category", "Bags"])
    summary[numeric_cols] = summary[numeric_cols].round(3)

    total_row = {col: None for col in summary.columns}
    total_row["Category"] = "TOTAL (by bag count)"
    total_row["Bags"] = int(summary["Bags"].sum())
    total_row["Avg Gross Loss %"] = round(
        (summary["Avg Gross Loss %"] * summary["Bags"]).sum() / summary["Bags"].sum(), 3
    )
    summary = pd.concat([summary, pd.DataFrame([total_row])], ignore_index=True)
    return summary


def _find_returned_recast_bags(state, recast_department="CAST"):
    from collections import defaultdict

    dept_visits = defaultdict(list)
    for row in state["stage_rows"]:
        if row["Department"] == recast_department:
            dept_visits[row["Bag No"]].append(row["Add Metal"])

    recast_bags = {}
    for bag_no, add_metal_values in dept_visits.items():
        if len(add_metal_values) > 1 and any(v > 0 for v in add_metal_values[1:]):
            recast_bags[bag_no] = len(add_metal_values)
    return recast_bags


def build_returned_bag_details(uploaded_files, metal_raw_name="Gold") -> pd.DataFrame:
    rows = []
    for uploaded_file in _as_file_list(uploaded_files):
        state = _load_workbook_state(uploaded_file, metal_raw_name)
        recast_bags = _find_returned_recast_bags(state)
        for bag_no, cast_visits in recast_bags.items():
            rows.append(
                {
                    "Source File": uploaded_file.name,
                    "Bag No": state["bag_label"][bag_no],
                    "Category": state["bag_data"][bag_no]["Category"],
                    "CAST Visits": cast_visits,
                }
            )
    return pd.DataFrame(rows)


def build_returned_bag_category_counts(returned_bag_details: pd.DataFrame) -> pd.DataFrame:
    if returned_bag_details.empty:
        return pd.DataFrame(columns=["Category", "Returned/Recast Bags", "% of Total Bags"])

    counts = (
        returned_bag_details.groupby("Category", dropna=False)
        .agg(**{"Returned/Recast Bags": ("Bag No", "count")})
        .reset_index()
        .sort_values("Returned/Recast Bags", ascending=False)
    )
    total = int(counts["Returned/Recast Bags"].sum())
    counts["% of Total Bags"] = (counts["Returned/Recast Bags"] / total * 100).round(2)
    return counts


def build_category_mismatch_summary(df: pd.DataFrame) -> pd.DataFrame:
    pivot = (
        df.pivot_table(index="Category", columns="Status", values="Bag No", aggfunc="count", fill_value=0)
        .reset_index()
    )
    for status in ("OK", "Mismatch", "No Data"):
        if status not in pivot.columns:
            pivot[status] = 0
    pivot["Total Bags"] = pivot[["OK", "Mismatch", "No Data"]].sum(axis=1)
    pivot["Mismatch %"] = (pivot["Mismatch"] / pivot["Total Bags"] * 100).round(2)
    pivot = pivot[["Category", "Total Bags", "OK", "Mismatch", "No Data", "Mismatch %"]]
    return pivot.sort_values("Mismatch", ascending=False)


def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Gross Loss %") -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        for col_idx, column in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
            cell.alignment = Alignment(horizontal="center")
            max_len = max([len(str(column))] + [len(str(v)) for v in df[column].astype(str).values[:200]])
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 28)
        worksheet.freeze_panes = "A2"
    return buffer.getvalue()
